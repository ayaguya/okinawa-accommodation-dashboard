#!/usr/bin/env python
# ============================================================
# transform_all_years.py
# ------------------------------------------------------------
# ・data/raw/excel/ 以下の h19shukuhaku.xls ～ r5 を対象
# ・和洋室の内訳は無視し、「計」の列を採用
# ・異なるシート（種別、規模、ホテル内訳）のデータを統合
# ・旧式の.xlsファイルと.xlsxファイルの両方に対応
# ============================================================

import sys
import unicodedata
import fnmatch
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import xlrd

# ---------- プロジェクト・パス ----------
BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

RAW_DIR = BASE_DIR / "data/raw/excel"
OUT_BY = BASE_DIR / "data/processed/by_year"
OUT_ALL = BASE_DIR / "data/processed/all"
OUT_BY.mkdir(parents=True, exist_ok=True)
OUT_ALL.mkdir(parents=True, exist_ok=True)

# ---------- 定数 ----------
try:
    from app.type_map import MUNICIPALITIES
except ImportError:
    print("Warning: app.type_map.MUNICIPALITIES not found. Using a dummy set.")
    MUNICIPALITIES = {"那覇市", "宜野湾市", "石垣市", "浦添市", "名護市", "糸満市", "沖縄市", "豊見城市", "うるま市", "宮古島市", "南城市", "国頭村", "大宜味村", "東村", "今帰仁村", "本部町", "恩納村", "宜野座村", "金武町", "読谷村", "嘉手納町", "北谷町", "北中城村", "中城村", "西原町", "与那原町", "南風原町", "八重瀬町", "多良間村", "竹富町", "与那国町"}

HEADER_ROWS = {
    (2018, 2023): 5, (2014, 2017): 4, (2007, 2013): 4,
}

YEAR_RULES = {
    (2018, 2023): {"s1": [r"市町村別・種別"], "s2": [r"規模別分類"], "s3": [r"ホテル内訳"], "s5": [r"住宅宿泊事業者|参考"]},
    (2014, 2017): {"s1": [r"市町村別・種別"], "s2": [r"規模別"], "s3": [r"ホテル内訳|種別・規模別"]},
    (2011, 2013): {"s1": [r"市町村別・種別"], "s2": [r"市町村別・規模別"]},
    (2007, 2010): {"s1": [r"全種別"], "s2": [r"ホテル・旅館"]},
}

JP2EN = {
    "市町村": "municipality", "市町村名": "municipality", "市区町村": "municipality",
    "宿泊施設数等": "", "宿泊種別": "accommodation_type", "宿泊施設種別": "accommodation_type", "種別": "accommodation_type",
    "軒数": "facilities", "施設数": "facilities", "施設数(軒)": "facilities", "宿泊施設数": "facilities",
    "客室数": "rooms", "客室数(室)": "rooms",
    "収容人数": "capacity", "収容人数(人)": "capacity",
    "届出軒数": "notifications", "住宅宿泊事業者届出軒数": "notifications",
    "ビジネス・宿泊特化型ホテル": "business_hotel", "ホテル・旅館": "hotel_ryokan",
    "民宿": "minshuku", "ペンション・貸別荘": "pension_villa", "ドミトリー・ゲストハウス": "dormitory_guesthouse",
    "ウィークリーマンション": "weekly_mansion", "団体経営施設": "group_facilities", "ユースホステル": "youth_hostel",
    "リゾートホテル": "resort_hotel", "シティーホテル": "city_hotel", "旅館": "ryokan",
    "大規模": "large", "中規模": "medium", "小規模": "small",
    "合 計": "total", "合計": "total",
    "平成": "h", "令和": "r",
}

ERA = {r"h(\d{1,2})": 1988, r"r(\d{1,2})": 2018}

# ---------- ヘルパ関数 ----------
def detect_year(name: str) -> int | None:
    n = name.lower()
    for patt, base in ERA.items():
        m = re.search(patt, n)
        if m: return base + int(m.group(1))
    m = re.search(r"(20\d{2})", n)
    return int(m.group(1)) if m else None

def header_rows_for(year: int) -> int:
    return next((n for (y0, y1), n in HEADER_ROWS.items() if y0 <= year <= y1), 4)

def rule_for(year: int) -> dict:
    return next((rule for (y0, y1), rule in YEAR_RULES.items() if y0 <= year <= y1), {})

def sheet_rows(sheet, start: int, end: int | None = None):
    if hasattr(sheet, 'iter_rows'):  # openpyxl sheet
        rows = sheet.iter_rows(min_row=start, max_row=end, values_only=False)
        return [[str(c.value).strip() if c.value is not None else "" for c in row] for row in rows]
    else:  # xlrd sheet
        end = end or sheet.nrows
        # xlrd is 0-indexed for rows and columns
        return [ [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)] for i in range(start - 1, end) ]

def find_sheet(key: str, year_rule: dict, sheet_names: list[str]):
    if key not in year_rule: return None
    for pat in year_rule[key]:
        rx = re.compile(pat, re.IGNORECASE)
        for n in sheet_names:
            if rx.search(n): return n
    return None

def first_muni_row(sheet, search_rows: int = 50) -> int | None:
    for idx, row in enumerate(sheet_rows(sheet, 1, search_rows), start=1):
        if any(unicodedata.normalize("NFKC", cell).strip() in MUNICIPALITIES for cell in row):
            return idx
    return None

def first_data_row(sheet, header_rows: int) -> int:
    muni_row = first_muni_row(sheet)
    if muni_row: return muni_row
    
    for i in range(header_rows + 1, header_rows + 20):
        try:
            row = sheet_rows(sheet, i, i)[0]
            if sum(1 for c in row if c and re.match(r'^-?\d+(\.\d+)?$', c)) >= 2:
                return i
        except IndexError: continue
    return header_rows + 1

def _clean_compare(val: str | None) -> str:
    if val is None: return ""
    s = unicodedata.normalize("NFKC", str(val))
    s = re.sub(r"[ \d\(\)①-⑩Ⅰ-Ⅹa-zA-Z]", "", s)
    return s.strip()

def ensure_municipality(df: pd.DataFrame) -> bool:
    if "municipality" in df.columns: return True
    if df.empty: return False

    for col_name in df.columns:
        # Check if a significant portion of the column contains municipality names
        if df[col_name].astype(str).str.contains('|'.join(MUNICIPALITIES)).sum() > 3:
            df.rename(columns={col_name: "municipality"}, inplace=True)
            return True
    return False

def flatten_headers(sheet, header_rows: int = 4) -> list[str | None]:
    rows = sheet_rows(sheet, 1, header_rows)
    
    # Forward fill header rows
    filled_rows = []
    for r in rows:
        last_val = None
        current_row = []
        for v in r:
            cleaned_v = str(v).strip()
            if cleaned_v and cleaned_v.lower() != 'none': last_val = cleaned_v
            current_row.append(last_val)
        filled_rows.append(current_row)

    final_headers = []
    sorted_jp2en = sorted(JP2EN.items(), key=lambda x: len(x[0]), reverse=True)
    known_metrics = {"facilities", "rooms", "capacity", "notifications"}

    for col_idx in range(len(filled_rows[0])):
        levels = [row[col_idx] for row in filled_rows if col_idx < len(row) and row[col_idx]]
        header_str = "_".join(dict.fromkeys(levels)) # Use unique levels
        
        # Explicitly identify municipality column
        if any(muni in header_str for muni in ["市町村", "市区町村"]):
            final_headers.append("municipality")
            continue

        # Skip irrelevant total columns (like '和'/'洋' totals)
        if any(s in header_str for s in ["和", "洋", "計", "合計"]) and not any(m in header_str for m in ["軒数", "客室数", "収容人数"]):
            final_headers.append(None)
            continue
            
        translated_str = header_str
        for jp, en in sorted_jp2en:
            translated_str = translated_str.replace(jp, en)

        # Clean the translated string
        cleaned_str = re.sub(r'[\(（][^)）]*[\)）]', '', translated_str) # Remove content in parens
        cleaned_str = re.sub(r'[^a-zA-Z0-9_]', '_', cleaned_str) # Replace non-alphanumeric with underscore
        cleaned_str = re.sub(r'_+', '_', cleaned_str).strip('_') # Consolidate underscores

        parts = [p for p in cleaned_str.split('_') if p and p != 'total']
        
        metric_part = next((p for p in parts if p in known_metrics), None)
        category_parts = [p for p in parts if p not in known_metrics and p != 'accommodation_type']
        
        final_name = None
        if metric_part:
            # Combine categories and metric for a full name
            final_name = "_".join(sorted(list(set(category_parts))) + [metric_part])
        elif 'accommodation_type' in parts:
             # Identify the accommodation_type column itself
             final_name = 'accommodation_type'
        
        final_headers.append(final_name)
    
    # Ensure unique headers by appending numbers if needed
    counts = {}
    unique_headers = []
    for h in final_headers:
        if h:
            if h in counts:
                unique_headers.append(f"{h}_{counts[h]}")
                counts[h] += 1
            else:
                counts[h] = 1
                unique_headers.append(h)
        else:
            unique_headers.append(None)
    return unique_headers

def process_sheet(ws, year, hr, id_vars, source):
    hdr = flatten_headers(ws, hr)
    data_start = first_data_row(ws, hr)
    df = pd.DataFrame(sheet_rows(ws, data_start), columns=hdr).dropna(how='all')

    if not ensure_municipality(df):
        print(f"  - No municipality column found. Skipping.")
        return None
    
    def normalize_muni_cell(name):
        if not isinstance(name, str): return None
        # Clean the name first
        name_cleaned = re.sub(r"[ 　\n※]|（再掲）", "", name)
        for m in MUNICIPALITIES:
            if m in name_cleaned:
                return m
        return None

    df['municipality'] = df['municipality'].apply(normalize_muni_cell)
    df = df.dropna(subset=['municipality']).copy()

    if df.empty:
        print(f"  - No valid municipality data after cleaning.")
        return None
    
    df["year"] = year
    for col, val in id_vars.items():
        if col not in df.columns:
            df[col] = val

    metrics = ("facilities", "rooms", "capacity", "notifications")
    value_cols = [c for c in df.columns if c and any(c.endswith(f"_{m}") or c == m for m in metrics)]

    for col in value_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    melted_df = df.melt(id_vars=["year", "municipality"] + list(id_vars.keys()), value_vars=value_cols, var_name="variable", value_name="value")
    if melted_df.empty or melted_df['value'].sum() == 0: return None

    melted_df['metric'] = melted_df['variable'].apply(lambda x: next((m for m in metrics if x.endswith(m)), None))
    melted_df = melted_df.dropna(subset=['metric'])

    def get_category(var_str, metric):
        base_name = var_str
        if base_name.endswith(f"_{metric}"):
            base_name = base_name[:-(len(metric) + 1)]
        elif base_name == metric:
            return "total" # It's a total if it's just the metric name
        return base_name

    melted_df['category'] = melted_df.apply(lambda row: get_category(row['variable'], row['metric']), axis=1)

    cat_col = next((k for k in id_vars if k != "accommodation_type"), None)
    if cat_col:
        melted_df[cat_col] = melted_df['category']
    elif 'accommodation_type' in id_vars:
        # Ffill for cases where type is listed once per group
        df['accommodation_type'] = df['accommodation_type'].ffill()
        melted_df['accommodation_type'] = melted_df['category'].replace('total', 'all_types')

    final_cols = ["year", "municipality", "accommodation_type", "scale_class", "hotel_subtype", "metric", "value"]
    df_long = melted_df.reindex(columns=final_cols)
    df_long['source'] = source
    
    return df_long

# ---------- メイン処理 ----------
all_frames = []

for xl in sorted(RAW_DIR.glob("*shukuhaku.*")):
    if fnmatch.fnmatch(xl.name, "~$*"): continue
    year = detect_year(xl.stem)
    if not year or year < 2007: continue

    print(f"\n=== Processing {xl.name} (Year: {year}) ===")
    
    try:
        if xl.suffix == '.xlsx':
            wb = load_workbook(xl, data_only=True, read_only=True)
            sheets = wb.sheetnames
            gs = lambda n: wb[n]
        elif xl.suffix == '.xls':
            wb = xlrd.open_workbook(xl, on_demand=True)
            sheets = wb.sheet_names()
            gs = lambda n: wb.sheet_by_name(n)
        else:
            continue
            
        yearly_dfs = {}
        yr_rule = rule_for(year)

        if (s1_name := find_sheet("s1", yr_rule, sheets)):
            print(f"  [s1] Found sheet: {s1_name}")
            yearly_dfs['s1'] = process_sheet(gs(s1_name), year, header_rows_for(year), {"accommodation_type": "N/A"}, "main_survey")
        if (s2_name := find_sheet("s2", yr_rule, sheets)):
            print(f"  [s2] Found sheet: {s2_name}")
            yearly_dfs['s2'] = process_sheet(gs(s2_name), year, header_rows_for(year), {"accommodation_type": "hotel_ryokan", "scale_class": "N/A"}, "main_survey")
        if (s3_name := find_sheet("s3", yr_rule, sheets)):
            print(f"  [s3] Found sheet: {s3_name}")
            yearly_dfs['s3'] = process_sheet(gs(s3_name), year, header_rows_for(year), {"accommodation_type": "hotel_ryokan", "hotel_subtype": "N/A"}, "main_survey")
        if (s5_name := find_sheet("s5", yr_rule, sheets)):
            print(f"  [s5] Found sheet: {s5_name}")
            if (res_df := process_sheet(gs(s5_name), year, header_rows_for(year), {"accommodation_type": "residential"}, "residential_act")):
                yearly_dfs['s5'] = res_df

        valid_dfs = [df for df in yearly_dfs.values() if df is not None]
        if not valid_dfs:
            print(f"  -> No data recognized for year {year}. SKIPPING.")
            continue

        yearly_df = pd.concat(valid_dfs, ignore_index=True)
        
        hotel_mask = yearly_df['accommodation_type'] == 'hotel_ryokan'
        df_hotels = yearly_df[hotel_mask]
        df_others = yearly_df[~hotel_mask]

        if not df_hotels.empty:
            df_hotels['detail_level'] = (df_hotels['hotel_subtype'].notna() & (df_hotels['hotel_subtype'] != "N/A")).astype(int) * 2 + \
                                        (df_hotels['scale_class'].notna() & (df_hotels['scale_class'] != "N/A")).astype(int)
            df_hotels = df_hotels.sort_values('detail_level', ascending=False)
            df_hotels = df_hotels.drop_duplicates(subset=['year', 'municipality', 'accommodation_type', 'metric'], keep='first')
        
        tidy = pd.concat([df_hotels, df_others], ignore_index=True).drop(columns=['detail_level'], errors='ignore')
        tidy = tidy[tidy['value'] > 0].reset_index(drop=True)

        final_cols = ["year", "municipality", "accommodation_type", "scale_class", "hotel_subtype", "metric", "value", "source"]
        tidy = tidy.reindex(columns=final_cols)
        
        tidy.to_csv(OUT_BY / f"accommodation_{year}.csv", index=False, encoding="utf-8-sig")
        print(f"  -> Saved accommodation_{year}.csv ({len(tidy):,} rows)")
        all_frames.append(tidy)

    except Exception as e:
        print(f"Error processing {xl.name}: {e}")
        import traceback
        traceback.print_exc()

# ---------- 全年度保存 ----------
if not all_frames:
    print("\nNo data generated for any year. ALL saved (0 rows).")
else:
    all_df = pd.concat(all_frames, ignore_index=True)
    all_df = all_df.sort_values(['year', 'municipality', 'accommodation_type']).reset_index(drop=True)
    all_df.to_csv(OUT_ALL / "okinawa_accommodation_tidy.csv", index=False, encoding="utf-8-sig")
    print(f"\n✅ ALL data processed and saved ({len(all_df):,} rows)")
