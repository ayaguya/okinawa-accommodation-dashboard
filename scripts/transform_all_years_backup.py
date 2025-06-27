#!/usr/bin/env python
# ============================================================
#  transform_all_years.py
# ------------------------------------------------------------
#  ・data/raw/excel/ 以下の h26shukuhaku.xlsx ～ r5 を対象
#  ・h25 以前は自動スキップ
#  ・和洋列は捨てて「計」列のみ採用
#  ・2024‑06 修正: dropna(errors="ignore") 削除／
#    「計」列名末尾の "_計"/"計" 除去／
#    規模別・ホテル内訳シートで empty 時も年度処理継続
# ============================================================

import sys
import unicodedata
import fnmatch
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook  # .xlsx

# ---------- プロジェクト・パス ----------
BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

RAW_DIR = BASE_DIR / "data/raw/excel"
OUT_BY = BASE_DIR / "data/processed/by_year"
OUT_ALL = BASE_DIR / "data/processed/all"
OUT_BY.mkdir(parents=True, exist_ok=True)
OUT_ALL.mkdir(parents=True, exist_ok=True)

# ---------- 定数 ----------
from app.type_map import MUNICIPALITIES  # 市町村正式名称のマスタ

HEADER_ROWS = {
    (2018, 2023): 5,  # H30～R5
    (2014, 2017): 4,  # H26～H29
    (2007, 2013): 4,  # H19～H25
}

YEAR_RULES = {
    (2018, 2023): {
        "s1": [r"市町村.*種別"],
        "s2": [r"規模別"],
        "s3": [r"ホテル内訳"],
        "s5": [r"住宅宿泊事業者|参考"],
    },
    (2014, 2017): {
        "s1": [r"市町村.*種別"],
        "s2": [r"規模別"],
        "s3": [r"ホテル内訳|種別・規模別"],
    },
    (2007, 2013): {
        "s1": [r"市町村.*種別|全種別"],
        "s2": [r"規模別|ホテル・旅館"],
    },
}

JP2EN = {
    "市町村": "municipality",
    "市町村名": "municipality",
    "市区町村": "municipality",
    "宿泊種別": "accommodation_type",
    "宿泊施設種別": "accommodation_type",
    "種別": "accommodation_type",
    "軒数": "facilities",
    "施設数": "facilities",
    "施設数(軒)": "facilities",
    "宿泊施設数": "facilities",
    "客室数": "rooms",
    "客室数(室)": "rooms",
    "収容人数": "capacity",
    "収容人数(人)": "capacity",
    "届出軒数": "notifications",
    "住宅宿泊事業者届出軒数": "notifications",
}

MIN_YEAR = 2014  # H26 未満はスキップ
ERA = {r"h(\d{1,2})": 1988, r"r(\d{1,2})": 2018}


# ---------- ヘルパ関数 ----------
def detect_year(name: str) -> int | None:
    n = name.lower()
    for patt, base in ERA.items():
        m = re.search(patt, n)
        if m:
            return base + int(m.group(1))
    m = re.search(r"(20\d{2})", n)
    return int(m.group(1)) if m else None


def header_rows_for(year: int) -> int:
    for (y0, y1), n in HEADER_ROWS.items():
        if y0 <= year <= y1:
            return n
    return 4


def rule_for(year: int) -> dict:
    for (y0, y1), rule in YEAR_RULES.items():
        if y0 <= year <= y1:
            return rule
    return {}


def sheet_rows(sheet, start: int, end: int | None = None):
    if hasattr(sheet, "iter_rows"):  # openpyxl
        return [[c.value for c in row] for row in sheet.iter_rows(min_row=start, max_row=end)]
    else:  # xlrd
        end = end or sheet.nrows
        return [sheet.row_values(i) for i in range(start - 1, end)]


def find_sheet(key: str, year_rule: dict, sheet_names: list[str]):
    if key not in year_rule:
        return None
    for pat in year_rule[key]:
        rx = re.compile(pat)
        for n in sheet_names:
            if rx.search(n):
                return n
    return None


def _normalize_name(s: str | None) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"^[\d０-９]+", "", s)
    s = re.sub(r"^[\(\（]?\d+[\)\）]?", "", s)
    return s.strip()


def first_muni_row(sheet, search_rows: int = 50) -> int | None:
    muni_set = set(MUNICIPALITIES)
    for idx, row in enumerate(sheet_rows(sheet, 1, search_rows), start=1):
        if any(_normalize_name(cell) in muni_set for cell in row):
            return idx
    return None


def first_data_row(sheet, header_rows: int) -> int:
    r = first_muni_row(sheet)
    if r:
        return r
    for i in range(header_rows + 1, header_rows + 11):
        row = sheet_rows(sheet, i, i)[0]
        if sum(isinstance(c, (int, float)) for c in row) >= 2:
            return i
    return header_rows + 1


def flatten_headers(sheet, header_rows: int = 4) -> list[str | None]:
         seen_muni = False
    rows = sheet_rows(sheet, 1, header_rows)
    # forward fill
    for r in rows:
        last = None
        for i, v in enumerate(r):
            last = v if v not in (None, "") else last
            r[i] = last
    flat = []
    muni_found = False 

    for col in zip(*rows):
        levels = [str(x).strip() if x else "" for x in col]
        if len(levels) < 4:
            levels += [""] * (4 - len(levels))
        a, b, c, d = levels[:4]

        # ① municipality 列（完全一致のみ）
        if any(k in a for k in ("市町村", "市区町村")):
            if not seen_muni:
                flat.append("municipality")   # 先頭だけ採用
                seen_muni = True
            else:
                flat.append(None)             # 以降は捨てる
            continue        h = "_".join(filter(None, [a, b, c]))
        for jp, en in JP2EN.items():
            if jp in h:
                h = h.replace(jp, en)
        h = re.sub(r"(facilities|rooms|capacity|notifications).*?$", r"\1", h)
        h = re.sub(r"_?計$", "", h)
        flat.append(h if any(t in h for t in ("facilities", "rooms", "capacity", "notifications")) else None)
    return flat


def _clean_compare(val: str | None) -> str:
    if val is None:
        return ""
    s = unicodedata.normalize("NFKC", str(val))
    s = re.sub(r"[ \d\(\)①-⑩Ⅰ-Ⅹa-zA-Z]", "", s)
    return s.strip()


def ensure_municipality(df: pd.DataFrame) -> bool:
    if "municipality" in df.columns:
        return True
    if len(df.columns) >= 3:
        df.rename(columns={df.columns[2]: "municipality"}, inplace=True)
        return True
    for col in df.columns:
        if df[col].dtype == object and df[col].apply(_clean_compare).isin(MUNICIPALITIES).sum() >= 5:
            df.rename(columns={col: "municipality"}, inplace=True)
            return True
    return False


def wide_to_long(df: pd.DataFrame, id_vars: list[str]) -> pd.DataFrame:
    id_vars = [c for c in id_vars if c in df.columns]
    metrics = ("facilities", "rooms", "capacity", "notifications")
    frames = []
    for m in metrics:
        cols = [c for c in df.columns if c and str(c).endswith(m)]
        if not cols:
            continue
        sub = df[id_vars + cols].copy()
        sub["metric"] = m
        sub["value"] = sub[cols].sum(axis=1, skipna=True)
        frames.append(sub[id_vars + ["metric", "value"]])
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def read_residential_sheet(ws, yr):
    hdr = flatten_headers(ws, header_rows_for(yr))
    start = first_muni_row(ws) or (header_rows_for(yr) + 1)
    df = pd.DataFrame(sheet_rows(ws, start), columns=hdr)
    if not ensure_municipality(df) or "notifications" not in df.columns:
        return None
    df = df.dropna(subset=["municipality"])
    df["municipality"] = df["municipality"].map(lambda x: str(x).replace("　", "").replace(" ", "").strip())
    df["year"] = yr
    return df[["year", "municipality", "notifications"]]


# ---------- メイン処理 ----------
all_frames: list[pd.DataFrame] = []

for xl in sorted(RAW_DIR.glob("*shukuhaku.*")):
    if fnmatch.fnmatch(xl.name, "~$*"):
        continue
    year = detect_year(xl.stem)
    if not year or year < MIN_YEAR:
        print(f"[skip] {xl.name}")
        continue

    print(f"=== {xl.name} ({year}) ===")
    wb = load_workbook(xl, data_only=True)
    sheets = wb.sheetnames
    gs = lambda n: wb[n]  # noqa: E731

    yearly: list[pd.DataFrame] = []
    yr_rule = rule_for(year)

    s1 = find_sheet("s1", yr_rule, sheets)
    s2 = find_sheet("s2", yr_rule, sheets)
    s3 = find_sheet("s3", yr_rule, sheets)
    s5 = find_sheet("s5", yr_rule, sheets)

    # -- 市町村別・種別 -------------------------------------------
    if s1:
        ws1 = gs(s1)
        hdr = flatten_headers(ws1, header_rows_for(year))
        data_start = first_data_row(ws1, header_rows_for(year))

        print("\n=== header check ===")
        for i, h in enumerate(hdr):
            print(f"{i:02d}: {repr(h)}")
        print("====================\n")
        data = sheet_rows(ws1, data_start) 
        df = pd.DataFrame(data, columns=hdr)
        print("metrics picked:",
              [c for c in df.columns if c and c.endswith(("facilities","rooms","capacity"))])

        start = first_data_row(ws1, header_rows_for(year))
        df = pd.DataFrame(sheet_rows(ws1, start), columns=hdr)
        if ensure_municipality(df):
            df = df.dropna(subset=["municipality"])
            if not df.empty:
                df["municipality"] = df["municipality"].str.replace("[ 　\n※]|（再掲）", "", regex=True)
                df["year"] = year
                yearly.append(
                    wide_to_long(df, ["year", "municipality", "accommodation_type"]).assign(
                        scale_class=pd.NA, hotel_subtype=pd.NA, source="main_survey"
                    )
                )

    # -- 規模別分類 ------------------------------------------------
    if s2:
        ws2 = gs(s2)
        hdr = flatten_headers(ws2, header_rows_for(year))
        start = first_data_row(ws2, header_rows_for(year))
        df = pd.DataFrame(sheet_rows(ws2, start), columns=hdr)
        if ensure_municipality(df):
            df = df.dropna(subset=["municipality"])
            if not df.empty:
                df["municipality"] = df["municipality"].str.replace("[ 　\n※]|（再掲）", "", regex=True)
                df["year"] = year
                for jp, en in {"大規模": "Large", "中規模": "Medium", "小規模": "Small"}.items():
                    cols = [c for c in df.columns if jp in str(c)]
                    if not cols:
                        continue
                    tmp = df.copy()
                    tmp["scale_class"] = en
                    res = wide_to_long(tmp, ["year", "municipality", "scale_class"])
                    if not res.empty:
                        yearly.append(res.assign(accommodation_type="ホテル・旅館", hotel_subtype=pd.NA, source="main_survey"))

    # -- ホテル内訳 ------------------------------------------------
    if s3:
        ws3 = gs(s3)
        hdr = flatten_headers(ws3, header_rows_for(year))
        start = first_data_row(ws3, header_rows_for(year))
        df = pd.DataFrame(sheet_rows(ws3, start), columns=hdr)
        if ensure_municipality(df):
            df = df.dropna(subset=["municipality"])
            if not df.empty:
                df["municipality"] = df["municipality"].str.replace("[ 　\n※]|（再掲）", "", regex=True)
                df["year"] = year
                for st in ["リゾートホテル", "ビジネス・宿泊特化型ホテル", "シティーホテル", "旅館"]:
                    cols = [c for c in df.columns if st in str(c)]
                    if not cols:
                        continue
                    tmp = df.copy()
                    tmp["hotel_subtype"] = st
                    yearly.append(
                        wide_to_long(tmp, ["year", "municipality", "hotel_subtype"]).assign(
                            accommodation_type="ホテル・旅館", scale_class=pd.NA, source="main_survey"
                        )
                    )

    # -- 住宅宿泊事業者 -------------------------------------------
    if s5 and year >= 2018:
        df = read_residential_sheet(gs(s5), year)
        if df is not None and not df.empty:
            yearly.append(
                df.rename(columns={"notifications": "value"}).assign(
                    accommodation_type=pd.NA,
                    scale_class=pd.NA,
                    hotel_subtype=pd.NA,
                    metric="notifications",
                    source="residential_act",
                )
            )

    # -- 年度保存 --------------------------------------------------
    if yearly:
        tidy = pd.concat(yearly, ignore_index=True)
        tidy.to_csv(OUT_BY / f"accommodation_{year}.csv", index=False, encoding="utf-8-sig")
        print(f"  -> saved accommodation_{year}.csv ({len(tidy):,} rows)")
        all_frames.append(tidy)
    else:
        print("  -> 認識できず SKIP")

# ---------- 全年度保存 ----------
if not all_frames:
    raise RuntimeError("生成データ 0 行")

all_df = pd.concat(all_frames, ignore_index=True)
all_df.to_csv(OUT_ALL / "okinawa_accommodation_tidy.csv", index=False, encoding="utf-8-sig")
print(f"\n✅ ALL saved ({len(all_df):,} rows)")
