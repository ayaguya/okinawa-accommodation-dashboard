"""
沖縄県宿泊施設データのExcelファイルをCSVに変換するスクリプト
"""
import sys
import os
from pathlib import Path

# プロジェクトルートをパスに追加
sys.path.append(str(Path(__file__).resolve().parent.parent))

import pandas as pd
import xlrd
from openpyxl import load_workbook

from app.type_map import (
    TYPE_MASTER, METRICS, YEARS, normalize_accommodation_type,
    normalize_metric, normalize_year, normalize_municipality
)

# ディレクトリ設定（プロジェクトルート基準）
BASE_DIR = Path(__file__).resolve().parent.parent
RAW_EXCEL_DIR = BASE_DIR / "data/raw/excel"
PROCESSED_DIR = BASE_DIR / "data/processed"

# Excelファイルの列マッピング
COLUMN_MAPPING = {
    # 市町村
    "市町村": "municipality",
    "市区町村": "municipality",
    "市町村名": "municipality",
    # 宿泊種別
    "宿泊種別": "accommodation_type",
    "種別": "accommodation_type",
    "宿泊施設種別": "accommodation_type",
    # 軒数
    "軒数": "facilities",
    "施設数": "facilities",
    "施設数(軒)": "facilities",
    # 客室数
    "客室数": "rooms",
    "客室数(室)": "rooms",
    # 収容人数
    "収容人数": "capacity",
    "収容人数(人)": "capacity",
    # 民泊届出数
    "民泊届出数": "minpaku_registrations",
    "民泊届出数(件)": "minpaku_registrations"
}

def detect_year(filename: str) -> int:
    """
    ファイル名から年度を検出
    
    Args:
        filename: ファイル名
        
    Returns:
        int: 西暦年
    """
    # ファイル名から年度文字列を抽出
    for year_str, year in YEARS.items():
        if year_str in filename:
            return year
    
    # 数値のみの場合も対応
    if filename.isdigit():
        return int(filename)
    
    # 見つからない場合は最小年度を返す
    return min(YEARS.values())

def normalize_column_name(col_name: str) -> str:
    """
    列名を正規化
    
    Args:
        col_name: 元の列名
        
    Returns:
        str: 正規化された列名
    """
    # 空白文字の削除と小文字化
    col_name = col_name.strip().lower()
    
    # マッピングに存在する場合はそのまま返す
    if col_name in COLUMN_MAPPING:
        return COLUMN_MAPPING[col_name]
        
    # 類似パターンのマッチング
    for pattern, mapped_col in COLUMN_MAPPING.items():
        if pattern.lower() in col_name:
            return mapped_col
            
    # 見つからない場合はNoneを返す
    return None

def read_xls_file(file_path: Path) -> pd.DataFrame:
    """
    .xlsファイルを読み込む
    
    Args:
        file_path: .xlsファイルのパス
        
    Returns:
        pd.DataFrame: 読み込んだデータ
    """
    try:
        # ワークブック読み込み
        workbook = xlrd.open_workbook(file_path)
        
        # 最適なシートの選択
        sheet = None
        for sheet_name in workbook.sheet_names():
            if any(keyword in sheet_name.lower() for keyword in ['データ', '統計', '集計']):
                sheet = workbook.sheet_by_name(sheet_name)
                break
        
        if not sheet:
            sheet = workbook.sheet_by_index(0)
            
        # データ読み込み
        data = []
        headers = []
        
        for row_idx in range(sheet.nrows):
            row = sheet.row(row_idx)
            
            # ヘッダー行の検出
            if not headers:
                headers = [normalize_column_name(str(cell.value)) for cell in row]
                if not any(headers):
                    continue
                
            # データ行の処理
            else:
                row_data = {}
                for cell_idx, header in enumerate(headers):
                    if header:
                        value = row[cell_idx].value
                        if isinstance(value, str):
                            value = value.strip()
                        elif isinstance(value, (int, float)):
                            value = float(value)
                        
                        if header == 'accommodation_type':
                            value = normalize_accommodation_type(value)
                        elif header == 'municipality':
                            value = normalize_municipality(value)
                        elif header in METRICS.values():
                            value = float(value) if value is not None else 0
                        
                        row_data[header] = value
                
                if row_data:
                    data.append(row_data)
        
        return pd.DataFrame(data)
        
    except Exception as e:
        print(f"エラー: {file_path.name} の読み込みに失敗しました: {str(e)}")
        return None

def read_xlsx_file(file_path: Path) -> pd.DataFrame:
    """
    .xlsxファイルを読み込み、データ開始行を自動検出

    Args:
        file_path: .xlsxファイルのパス

    Returns:
        pd.DataFrame: 読み込んだデータ
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = None

        # 最適なシート名を選択
        for sheet_name in workbook.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['データ', '統計', '集計', '市町村']):
                sheet = workbook[sheet_name]
                break
        if not sheet:
            sheet = workbook.active

        # データ開始行を自動検出（例: 那覇市など既知の市町村が現れる最初の行）
        known_municipalities = ['那覇市', '石垣市', '宮古島市', '名護市']
        start_row = 1
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=30), start=1):
            for cell in row:
                if isinstance(cell.value, str) and any(name in cell.value for name in known_municipalities):
                    start_row = i
                    break
            if start_row != 1:
                break

        headers = []
        data = []

        for i, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
            if not headers:
                headers = [normalize_column_name(str(cell.value) if cell.value else "") for cell in row]
                continue

            row_data = {}
            for cell, header in zip(row, headers):
                if not header:
                    continue
                value = cell.value
                if isinstance(value, str):
                    value = value.strip()
                elif isinstance(value, (int, float)):
                    value = float(value)
                if header == 'accommodation_type':
                    value = normalize_accommodation_type(value)
                elif header == 'municipality':
                    value = normalize_municipality(value)
                elif header in METRICS.values():
                    value = float(value) if value is not None else 0
                row_data[header] = value

            if row_data:
                data.append(row_data)

        return pd.DataFrame(data)

    except Exception as e:
        print(f"エラー: {file_path.name} の読み込みに失敗しました: {str(e)}")
        return None

def convert_excel_to_csv(excel_file: Path) -> None:
    """
    単一のExcelファイルをCSVに変換
    
    Args:
        excel_file: Excelファイルのパス
    """
    try:
        # ファイルの種類に応じて読み込み
        if excel_file.suffix == '.xls':
            df = read_xls_file(excel_file)
        else:  # .xlsx
            df = read_xlsx_file(excel_file)
            
        if df is None or df.empty:
            print(f"警告: {excel_file.name} からデータを抽出できませんでした")
            return
            
        # 年度追加
        year = detect_year(excel_file.stem)
        df['year'] = year
        
        # データ保存
        output_file = PROCESSED_DIR / f"accommodation_data_{year}.csv"
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"変換完了: {excel_file.name} → {output_file.name}")
        
    except Exception as e:
        print(f"エラー: {excel_file.name} の変換に失敗しました: {str(e)}")

def main():
    """メイン関数"""
    print("=== 沖縄県宿泊施設データ変換スクリプト ===")
    
    # 出力ディレクトリ作成
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    
    # Excelファイル検索
    excel_files = list(RAW_EXCEL_DIR.glob("*.xlsx")) + list(RAW_EXCEL_DIR.glob("*.xls"))
    
    if not excel_files:
        print("エラー: Excelファイルが見つかりません。data/raw/excel/ にファイルを配置してください。")
        return
    
    # ファイル数表示
    print(f"\n見つかったExcelファイル数: {len(excel_files)}")
    for file in excel_files:
        print(f"- {file.name}")
    
    # 変換実行
    print("\n変換を開始します...")
    for excel_file in excel_files:
        convert_excel_to_csv(excel_file)
    
    print("\n変換処理が完了しました！")

if __name__ == "__main__":
    main()
