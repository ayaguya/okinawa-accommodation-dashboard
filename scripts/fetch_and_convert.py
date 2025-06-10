import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import shutil

def download_excel_files():
    """
    沖縄県観光統計調査のExcelファイルをダウンロード
    """
    base_url = "https://www.pref.okinawa.jp"
    url = f"{base_url}/site/kankou/1000000000000/1000000000000/2000000000000.html"
    
    # ページのHTMLを取得
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Excelファイルのリンクを検索
    excel_links = []
    for link in soup.find_all('a', href=True):
        href = link['href']
        if href.endswith('.xlsx') or href.endswith('.xls'):
            excel_links.append(href)
    
    # Excelファイルをダウンロード
    for link in excel_links:
        file_url = f"{base_url}{link}"
        response = requests.get(file_url)
        
        # ファイル名を取得
        file_name = os.path.basename(link)
        file_path = os.path.join('data', 'raw', file_name)
        
        # ファイルを保存
        with open(file_path, 'wb') as f:
            f.write(response.content)
            print(f"Downloaded: {file_name}")

def convert_excel_to_csv():
    """
    ダウンロードされたExcelファイルをCSVに変換
    """
    raw_dir = 'data/raw'
    
    # rawディレクトリ内のExcelファイルを処理
    for file_name in os.listdir(raw_dir):
        if file_name.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(raw_dir, file_name)
            
            try:
                # Excelファイルを読み込み
                wb = load_workbook(file_path, read_only=True)
                
                # 各シートを処理
                for sheet_name in wb.sheetnames:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # ファイル名を生成
                    csv_name = f"{os.path.splitext(file_name)[0]}_{sheet_name}.csv"
                    csv_path = os.path.join(raw_dir, csv_name)
                    
                    # CSVに保存
                    df.to_csv(csv_path, index=False, encoding='utf-8')
                    print(f"Converted: {csv_name}")
                    
            except Exception as e:
                print(f"Error processing {file_name}: {str(e)}")

def main():
    # データディレクトリの作成
    os.makedirs('data/raw', exist_ok=True)
    
    # Excelファイルのダウンロード
    download_excel_files()
    
    # ExcelファイルをCSVに変換
    convert_excel_to_csv()

if __name__ == "__main__":
    main()
